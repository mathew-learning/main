import csv
import openpyxl


def xlsx_search(file, header, term):
    # Returns the row in a list.
    try:
        wrkdk = openpyxl.load_workbook(file)
    except FileNotFoundError :
        print("File not found")
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid format")
    else:
        sh = wrkdk.active
        for row in sh.iter_rows(min_row=1, max_row=sh.max_row, values_only=True):
            for col in sh.iter_cols(min_row=1, max_row=sh.max_column, values_only=True):
                if str(header) in col and str(term) in row:
                    row_values = list(row)
                    print(row_values)


xlsx_search("sample.xlsx", "RULE_STATUS", "NOK")
