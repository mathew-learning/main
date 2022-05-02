import csv
import openpyxl
import pandas as pd

wrkdk = openpyxl.load_workbook('sample.xlsx')
sh = wrkdk.active
for row in sh.iter_rows(min_row=1, max_row=sh.max_row, values_only=True):
    for col in sh.iter_cols(min_row=1, max_row=10, values_only=True):
        if "RULE_STATUS" in col and "NOK" in row:
            print(list(row))
# print(list(sh.columns))
# print(sh.max_row)
# print(sh.max_column)
#
# row_list = []
# for i in range(0, sh.max_row):
#     for obj in list(sh.rows)[i]:
#         row_list.append(obj.value)
#     for j in range(sh.max_column):
#         for obj in list(sh.columns)[j]:
#             # print(obj.value)
#             row_list.append[i](obj.value)
# print(row_list)
# for i in range(0, sh.max_row):
#     for obj in list(sh.rows)[i]:
#         row_list.append(obj.value)
# # print(row_list)
# # print(range(sh.max_row))
# for j in range(sh.max_column):
#     for obj in list(sh.columns)[j]:
#         # print(obj.value)
#         row_list.append(obj.value)
# # print(row_list)
#
#
# listr_ = list(sh.rows)
# lsitc_ = list(sh.columns)
# dict_r = {}
# list_re = []
# list_ce = []
# for obj in listr_[1]:
#     if obj.value is None:
#         continue
#     list_ce.append(obj.value)
#     dict_r[obj.value] = None
# # for obj in lsitc_[0]:
# #     list_re.append(obj.value)
# # print(list_re)
# # print(list_ce)
#     # print(obj.value)
#
#
# # print(dict_r)
# read_file = pd.read_excel('sample.xlsx')
# read_file.to_csv('sample.csv')


# with open('sample.csv', newline='') as csvfile:
#     reader = csv.reader(csvfile)
#     for row in reader:
#         if 'NOK' in row:
#             print(', '.join(row))


# with open('sample.csv', newline='') as csvfile:
#     fieldnames = list_ce
#     reader = csv.DictReader(csvfile, fieldnames=fieldnames)
#     for row in reader:
#         print(row)
